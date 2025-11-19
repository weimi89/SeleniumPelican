#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
WEDI 運費未請款明細下載工具
使用基礎類別實作運費未請款明細查詢功能
直接抓取HTML表格並轉換為Excel檔案
"""

import argparse
import re
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from bs4 import BeautifulSoup, Tag
from selenium.webdriver.common.by import By

from src.core.constants import Timeouts
from src.core.improved_base_scraper import ImprovedBaseScraper
from src.core.multi_account_manager import MultiAccountManager
from src.utils.windows_encoding_utils import check_pythonunbuffered

# 使用共用的模組和改進版基礎類別

# 檢查 PYTHONUNBUFFERED 環境變數
check_pythonunbuffered()


class UnpaidScraper(ImprovedBaseScraper):
    """
    WEDI 運費未請款明細下載工具
    繼承 BaseScraper 實作運費未請款明細查詢
    """

    def __init__(
        self, username, password, headless=False, download_base_dir="downloads"
    ):
        # 構建 URL
        url = "http://wedinlb03.e-can.com.tw/wEDI2012/wedilogin.asp"

        # 設定此爬蟲要使用的環境變數 key
        self.download_dir_env_key = "UNPAID_DOWNLOAD_WORK_DIR"
        self.download_ok_dir_env_key = "UNPAID_DOWNLOAD_OK_DIR"

        # 調用父類構造函數
        super().__init__(
            url=url, username=username, password=password, headless=headless
        )

        # download_base_dir 保留以保持向後相容，但標註為已棄用
        self.download_base_dir = download_base_dir  # Deprecated: 改用環境變數 UNPAID_DOWNLOAD_WORK_DIR

        # 設定結束時間為當日
        self.end_date = datetime.now().strftime("%Y%m%d")

    def navigate_to_unpaid_freight_page(self) -> bool:
        """導航到運費未請款明細頁面"""
        assert self.driver is not None, "WebDriver must be initialized"
        self.logger.info("🧭 導航至運費未請款明細頁面...")

        try:
            # 已經在 datamain iframe 中（由 ImprovedBaseScraper.navigate_to_query() 切換），等待頁面載入
            time.sleep(Timeouts.IFRAME_SWITCH)

            # 搜尋所有連結，找出運費未請款相關項目
            all_links = self.driver.find_elements(By.TAG_NAME, "a")
            self.logger.info(f"   找到 {len(all_links)} 個連結", links_count=len(all_links))

            unpaid_freight_link = None
            for i, link in enumerate(all_links):
                try:
                    link_text = link.text.strip()
                    if link_text:
                        # 檢查運費未請款明細相關關鍵字
                        if (
                            ("運費" in link_text and "未請款" in link_text)
                            or ("未請款" in link_text and "明細" in link_text)
                            or (
                                "運費" in link_text
                                and "明細" in link_text
                                and "請款" in link_text
                            )
                        ):
                            unpaid_freight_link = link
                            self.logger.info(
                                f"   ✅ 找到運費未請款明細連結: {link_text}",
                                link_text=link_text,
                                match_type="unpaid_freight",
                            )
                            break
                        elif "運費" in link_text and "明細" in link_text:
                            self.logger.debug(
                                f"   🔍 找到運費相關連結: {link_text}",
                                link_text=link_text,
                                match_type="freight_related",
                            )
                except Exception:
                    continue

            if unpaid_freight_link:
                # 使用JavaScript點擊避免元素遮蔽問題
                self.driver.execute_script("arguments[0].click();", unpaid_freight_link)
                time.sleep(Timeouts.PAGE_LOAD)
                self.logger.log_operation_success("點擊運費未請款明細連結")
                return True
            else:
                self.logger.warning("❌ 未找到運費未請款明細連結")
                # 嘗試驗證頁面是否包含運費未請款功能
                page_text = self.driver.page_source
                if "運費" in page_text and ("未請款" in page_text or "明細" in page_text):
                    self.logger.info(
                        "✅ 頁面包含運費未請款功能，繼續流程", fallback_method="page_content_check"
                    )
                    return True
                else:
                    self.logger.error("❌ 頁面不包含運費未請款功能", page_check="failed")
                    return False

        except Exception as e:
            # 使用診斷管理器捕獲導航異常
            diagnostic_report = self.diagnostic_manager.capture_exception(
                e,
                context={
                    "operation": "navigate_to_unpaid_freight_page",
                    "username": self.username,
                    "current_url": self.driver.current_url if self.driver else None,
                    "links_found": len(self.driver.find_elements(By.TAG_NAME, "a"))
                    if self.driver
                    else 0,
                },
                capture_screenshot=True,
                capture_page_source=True,
                driver=self.driver,
            )

            self.logger.log_operation_failure(
                "導航到運費未請款明細頁面", e, diagnostic_report=diagnostic_report
            )
            return False

    def set_end_date(self) -> bool:
        """設定結束時間為當日 - 不需要使用者輸入"""
        assert self.driver is not None, "WebDriver must be initialized"
        self.logger.info("📅 設定結束時間為當日...")

        self.logger.info(f"📅 結束時間: {self.end_date}", end_date=self.end_date)

        try:
            # 已經在iframe中，嘗試尋找日期輸入框
            date_inputs = self.driver.find_elements(
                By.CSS_SELECTOR, 'input[type="text"]'
            )

            if len(date_inputs) >= 1:
                try:
                    # 填入結束時間（當日）
                    # 通常運費未請款明細只需要一個結束時間
                    date_inputs[-1].clear()  # 使用最後一個輸入框作為結束時間
                    date_inputs[-1].send_keys(self.end_date)
                    self.logger.log_operation_success("設定結束時間", end_date=self.end_date)
                except Exception as date_error:
                    self.logger.warning(
                        f"⚠️ 填入結束時間失敗: {date_error}", error=str(date_error)
                    )

                # 嘗試點擊查詢按鈕
                query_button_found = False
                button_selectors = [
                    'input[value*="查詢"]',
                    'button[title*="查詢"]',
                    'input[type="submit"]',
                    'button[type="submit"]',
                    'input[value*="搜尋"]',
                ]

                for selector in button_selectors:
                    try:
                        query_button = self.driver.find_element(
                            By.CSS_SELECTOR, selector
                        )
                        query_button.click()
                        self.logger.log_operation_success("點擊查詢按鈕", selector=selector)
                        time.sleep(Timeouts.QUERY_SUBMIT)
                        query_button_found = True
                        break
                    except Exception:
                        continue

                if not query_button_found:
                    self.logger.warning("⚠️ 未找到查詢按鈕，直接繼續流程")
            else:
                self.logger.warning("⚠️ 未找到日期輸入框，可能不需要設定日期")

            return True

        except Exception as e:
            # 使用診斷管理器捕獲日期設定異常
            diagnostic_report = self.diagnostic_manager.capture_exception(
                e,
                context={
                    "operation": "set_end_date",
                    "username": self.username,
                    "end_date": self.end_date,
                    "current_url": self.driver.current_url if self.driver else None,
                },
                capture_screenshot=True,
                capture_page_source=True,
                driver=self.driver,
            )

            self.logger.warning(
                f"⚠️ 結束時間設定過程中出現問題，但繼續執行: {e}",
                error=str(e),
                continue_execution=True,
                diagnostic_report=diagnostic_report,
            )
            return True  # 即使失敗也返回True，讓流程繼續

    def extract_table_data_to_excel(self) -> Optional[str]:
        """直接從HTML表格提取數據並轉換為Excel檔案"""
        assert self.driver is not None, "WebDriver must be initialized"
        self.logger.info("📊 提取表格數據並轉換為Excel...")

        try:
            # 等待頁面完全載入
            time.sleep(Timeouts.PAGE_LOAD)

            # 獲取頁面HTML
            page_html = self.driver.page_source
            soup = BeautifulSoup(page_html, "html.parser")

            # 尋找包含數據的表格
            tables = soup.find_all("table")
            self.logger.info(f"   找到 {len(tables)} 個表格", tables_count=len(tables))

            main_table: Optional[Tag] = None
            max_rows = 0

            # 找到最大的表格（通常是包含數據的主表格）
            for table in tables:
                if not isinstance(table, Tag):
                    continue
                rows = table.find_all("tr")
                if len(rows) > max_rows:
                    max_rows = len(rows)
                    main_table = table

            if not main_table or max_rows < 2:  # 至少要有表頭和一行數據
                self.logger.error("❌ 未找到包含數據的表格", max_rows=max_rows)
                return None

            self.logger.info(f"✅ 找到主要數據表格，共 {max_rows} 行", table_rows=max_rows)

            # 提取表格數據
            table_data: List[List[str]] = []
            rows = main_table.find_all("tr")

            for row_index, row in enumerate(rows):
                if not isinstance(row, Tag):
                    continue
                row_data: List[str] = []
                cells = row.find_all(["td", "th"])

                for cell in cells:
                    # 清理單元格內容
                    cell_text = cell.get_text(strip=True)
                    # 移除HTML實體和多餘空白
                    cell_text = cell_text.replace("\u00a0", " ").replace("\xa0", " ")
                    cell_text = re.sub(r"\s+", " ", cell_text).strip()
                    row_data.append(cell_text)

                if row_data:  # 只添加非空行
                    table_data.append(row_data)
                    if row_index < 5:  # 只顯示前5行的內容用於調試
                        self.logger.debug(
                            f"   行 {row_index + 1}: {row_data[:5]}...",
                            row_index=row_index + 1,
                            row_preview=row_data[:5],
                        )

            if not table_data:
                self.logger.error("❌ 表格中沒有找到數據")
                return None

            self.logger.log_data_info("成功提取表格數據", count=len(table_data))

            # 創建Excel檔案
            wb = openpyxl.Workbook()
            ws = wb.active
            assert ws is not None, "Workbook active sheet should not be None"
            ws.title = "運費未請款明細"

            # 將數據寫入工作表
            for row_index, row_data in enumerate(table_data, 1):
                for col_index, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_index, column=col_index, value=cell_value)

            # 設定表頭樣式（如果有表頭）
            if len(table_data) > 0:
                from openpyxl.styles import Font, PatternFill

                # 第一行設為表頭樣式
                for col_index in range(1, len(table_data[0]) + 1):
                    cell = ws.cell(row=1, column=col_index)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )

            # 自動調整欄寬
            from openpyxl.cell.cell import Cell

            for column in ws.columns:
                max_length = 0
                # 取得第一個 Cell 的 column_letter (跳過 MergedCell)
                column_letter = None
                for cell in column:
                    if isinstance(cell, Cell) and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                        break

                if column_letter is None:
                    continue

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大寬度限制
                ws.column_dimensions[column_letter].width = adjusted_width

            # 生成檔案名稱：運費未請款明細_{帳號}_{結束時間}.xlsx
            filename = f"運費未請款明細_{self.username}_{self.end_date}.xlsx"
            
            # 檢查檔案是否已下載
            exists, existing_path = self.is_file_downloaded(filename)
            if exists:
                wb.close()  # 關閉未使用的 workbook
                self.logger.info(
                    f"⏭️ 檔案已存在，跳過生成: {filename}",
                    location=str(existing_path)
                )
                return str(existing_path)
            
            file_path = self.download_dir / filename

            # 確保下載目錄存在且可寫入（提供詳細診斷訊息）
            self.ensure_directory_writable(self.download_dir)

            # 保存檔案
            wb.save(file_path)
            wb.close()

            file_size = file_path.stat().st_size
            rows_count = len(table_data)
            cols_count = len(table_data[0]) if table_data else 0

            self.logger.log_operation_success(
                "生成運費未請款明細Excel",
                filename=filename,
                file_size_bytes=file_size,
                rows_count=rows_count,
                cols_count=cols_count,
            )

            return str(file_path)

        except Exception as e:
            # 使用診斷管理器捕獲表格提取異常
            diagnostic_report = self.diagnostic_manager.capture_exception(
                e,
                context={
                    "operation": "extract_table_data_to_excel",
                    "username": self.username,
                    "end_date": self.end_date,
                    "download_dir": str(self.download_dir),
                    "current_url": self.driver.current_url if self.driver else None,
                    "page_source_available": bool(
                        self.driver and hasattr(self.driver, "page_source")
                    ),
                },
                capture_screenshot=True,
                capture_page_source=True,
                driver=self.driver,
            )

            self.logger.log_operation_failure(
                "提取表格數據",
                e,
                diagnostic_report=diagnostic_report,
                username=self.username,
                end_date=self.end_date,
            )
            return None

    def run_full_process(self) -> List[str]:
        """執行完整的自動化流程"""
        all_downloads: List[str] = []

        try:
            self.logger.info("=" * 60)
            self.logger.info(
                f"📊 開始執行 WEDI 運費未請款明細下載流程 (帳號: {self.username})",
                username=self.username,
                process="unpaid_freight",
            )
            self.logger.info("=" * 60)

            # 1. 瀏覽器已在 __init__ 中初始化

            # 2. 登入
            login_success = self.login()
            if not login_success:
                self.logger.log_operation_failure("登入", Exception("登入失敗"))
                return []

            # 3. 導航到查詢頁面（基礎導航）
            nav_success = self.navigate_to_query()
            if not nav_success:
                self.logger.log_operation_failure("基礎導航", Exception("導航失敗"))
                return []

            # 4. 導航到運費未請款明細頁面
            unpaid_nav_success = self.navigate_to_unpaid_freight_page()
            if not unpaid_nav_success:
                self.logger.log_operation_failure("運費未請款明細頁面導航", Exception("頁面導航失敗"))
                return []

            # 5. 設定結束時間（當日）
            self.set_end_date()

            # 6. 直接提取表格數據並轉換為Excel
            excel_file = self.extract_table_data_to_excel()

            if excel_file:
                all_downloads.append(excel_file)
                self.logger.log_operation_success(
                    "運費未請款明細下載完成",
                    username=self.username,
                    files_downloaded=len(all_downloads),
                )
            else:
                self.logger.warning(
                    f"⚠️ 帳號 {self.username} 沒有找到可下載的數據",
                    username=self.username,
                    status="no_data",
                )

            return all_downloads

        except Exception as e:
            self.logger.log_operation_failure("流程執行", e)
            return all_downloads
        finally:
            self.close()


def main():
    """主程式入口"""
    from src.core.logging_config import get_logger

    logger = get_logger("unpaid_scraper_main")

    parser = argparse.ArgumentParser(description="WEDI 運費未請款明細下載工具")
    parser.add_argument("--headless", action="store_true", help="使用無頭模式")

    args = parser.parse_args()

    # 顯示結束時間（當日）
    end_date = datetime.now().strftime("%Y%m%d")
    logger.info(f"📅 結束時間: {end_date} (當日)", end_date=end_date)

    try:
        # 使用多帳號管理器
        logger.info("📊 WEDI 運費未請款明細下載工具")

        manager = MultiAccountManager("accounts.json")
        manager.run_all_accounts(
            scraper_class=UnpaidScraper,
            headless_override=args.headless if args.headless else None,
        )

        return 0

    except (FileNotFoundError, ValueError, RuntimeError) as e:
        logger.error(f"⛔ 錯誤: {e}", error=str(e), error_type=type(e).__name__)
        return 1
    except KeyboardInterrupt:
        logger.error("⛔ 使用者中斷執行", error_type="KeyboardInterrupt")
        return 1
    except Exception as e:
        logger.error(
            f"⛔ 未知錯誤: {e}", error=str(e), error_type=type(e).__name__, exc_info=True
        )
        return 1


if __name__ == "__main__":
    main()
