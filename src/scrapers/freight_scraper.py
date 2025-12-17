#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
WEDI 運費查詢工具
使用基礎類別實作運費(月結)結帳資料查詢功能
"""

import argparse
import json
import time
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

import openpyxl
from selenium.webdriver.common.by import By

from src.core.constants import Timeouts
from src.core.improved_base_scraper import ImprovedBaseScraper
from src.core.multi_account_manager import MultiAccountManager
from src.core.type_aliases import DownloadResult
from src.utils.windows_encoding_utils import check_pythonunbuffered

# 使用共用的模組和改進版基礎類別

# 檢查 PYTHONUNBUFFERED 環境變數
check_pythonunbuffered()


class FreightScraper(ImprovedBaseScraper):
    """
    WEDI 運費查詢工具
    繼承 BaseScraper 實作運費(月結)結帳資料查詢
    """

    def __init__(
        self,
        username,
        password,
        headless=False,
        download_base_dir="downloads",
        start_month=None,
        end_month=None,
    ):
        # 構建 URL
        url = "http://wedinlb03.e-can.com.tw/wEDI2012/wedilogin.asp"

        # 設定此爬蟲要使用的環境變數 key
        self.download_dir_env_key = "FREIGHT_DOWNLOAD_WORK_DIR"
        self.download_ok_dir_env_key = "FREIGHT_DOWNLOAD_OK_DIR"

        # 調用父類構造函數
        super().__init__(
            url=url, username=username, password=password, headless=headless
        )

        # 子類特有的屬性
        self.start_month = start_month
        self.end_month = end_month
        # download_base_dir 保留以保持向後相容，但標註為已棄用
        self.download_base_dir = download_base_dir  # Deprecated: 改用環境變數 FREIGHT_DOWNLOAD_WORK_DIR

        # 轉換月份為日期格式供日期操作使用
        self.start_date = None
        self.end_date = None

        if start_month:
            try:
                year = int(start_month[:4])
                month = int(start_month[4:])
                self.start_date = datetime(year, month, 1)
            except (ValueError, IndexError):
                self.logger.error(f"❌ 月份格式錯誤: {start_month}")

        if end_month:
            try:
                year = int(end_month[:4])
                month = int(end_month[4:])
                # 計算該月最後一天
                if month == 12:
                    next_year = year + 1
                    next_month = 1
                else:
                    next_year = year
                    next_month = month + 1
                last_day = datetime(next_year, next_month, 1) - timedelta(days=1)
                self.end_date = last_day
            except (ValueError, IndexError):
                self.logger.error(f"❌ 月份格式錯誤: {end_month}")

    def get_default_date_range(self):
        """獲取預設月份範圍：上個月"""
        today = datetime.now()

        # 計算上個月
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year

        prev_month_str = f"{prev_year:04d}{prev_month:02d}"

        # 預設開始和結束月份都是上個月
        return prev_month_str, prev_month_str

    def navigate_to_freight_page(self) -> bool:
        """導航到運費查詢頁面 (2-7)運費(月結)結帳資料查詢"""
        assert self.driver is not None, "WebDriver must be initialized"
        self.logger.info(f"🧭 導航至運費查詢頁面...", operation="search")

        try:
            # 已經在 datamain iframe 中（由 ImprovedBaseScraper.navigate_to_query() 切換），等待頁面載入
            time.sleep(Timeouts.IFRAME_SWITCH)

            # 搜尋所有連結，找出運費相關項目
            all_links = self.driver.find_elements(By.TAG_NAME, "a")
            self.logger.info(f"   找到 {len(all_links)} 個連結")

            freight_link = None
            for i, link in enumerate(all_links):
                try:
                    link_text = link.text.strip()
                    if link_text:
                        # 檢查運費(月結)結帳資料查詢相關關鍵字
                        if (
                            ("運費" in link_text and "月結" in link_text)
                            or ("2-7" in link_text and "運費" in link_text)
                            or ("結帳資料" in link_text and "運費" in link_text)
                        ):
                            freight_link = link
                            self.logger.info(
                                f"   ✅ 找到運費查詢連結: {link_text}", operation="search"
                            )
                            break
                        elif "運費" in link_text:
                            self.logger.debug(f"   🔍 找到運費相關連結: {link_text}")
                except Exception:
                    continue

            if freight_link:
                # 使用JavaScript點擊避免元素遮蔽問題
                self.driver.execute_script("arguments[0].click();", freight_link)
                time.sleep(Timeouts.PAGE_LOAD)
                self.logger.info(f"✅ 已點擊運費查詢連結", operation="search")
                return True
            else:
                self.logger.log_operation_failure("運費查詢連結搜尋", "未找到運費查詢連結")
                # 嘗試驗證頁面是否包含運費功能
                page_text = self.driver.page_source
                if "運費" in page_text and ("月結" in page_text or "結帳" in page_text):
                    self.logger.info(f"✅ 頁面包含運費查詢功能，繼續流程", operation="search")
                    return True
                else:
                    self.logger.log_operation_failure("運費查詢功能驗證", "頁面不包含運費查詢功能")
                    return False

        except Exception as e:
            # 使用診斷管理器捕獲導航異常
            diagnostic_report = self.diagnostic_manager.capture_exception(
                e,
                context={
                    "operation": "navigate_to_freight_page",
                    "username": self.username,
                    "current_url": self.driver.current_url if self.driver else None,
                    "links_found": len(self.driver.find_elements(By.TAG_NAME, "a"))
                    if self.driver
                    else 0,
                },
                capture_screenshot=True,
                capture_page_source=True,
                driver=self.driver,
            )

            self.logger.log_operation_failure(
                "導航到運費查詢頁面", e, diagnostic_report=diagnostic_report
            )
            return False

    def set_date_range(self) -> bool:
        """設定查詢月份範圍 - 基於wedi_selenium_scraper.py的邏輯但適配月份"""
        assert self.driver is not None, "WebDriver must be initialized"
        self.logger.info(f"📅 設定月份範圍...", operation="config")

        # 使用指定的月份範圍，如果沒有指定則使用預設值
        if self.start_month and self.end_month:
            start_month = self.start_month
            end_month = self.end_month
        else:
            # 使用預設值（上個月）
            start_month_str, end_month_str = self.get_default_date_range()
            start_month = start_month_str
            end_month = end_month_str

        self.logger.info(f"📅 查詢月份範圍: {start_month} ~ {end_month}", operation="search")

        try:
            # 已經在iframe中，嘗試尋找月份輸入框
            date_inputs = self.driver.find_elements(
                By.CSS_SELECTOR, 'input[type="text"]'
            )

            if len(date_inputs) >= 2:
                try:
                    # 填入開始月份
                    date_inputs[0].clear()
                    date_inputs[0].send_keys(start_month)
                    self.logger.info(f"✅ 已設定開始月份: {start_month}", operation="config")

                    # 填入結束月份
                    date_inputs[1].clear()
                    date_inputs[1].send_keys(end_month)
                    self.logger.info(f"✅ 已設定結束月份: {end_month}", operation="config")
                except Exception as date_error:
                    self.logger.error(f"⚠️ 填入月份失敗: {date_error}", error=str(date_error))

                # 嘗試點擊查詢按鈕（與wedi_selenium_scraper.py相同的邏輯）
                query_button_found = False
                button_selectors = [
                    'input[value*="查詢"]',
                    'button[title*="查詢"]',
                    'input[type="submit"]',
                    'button[type="submit"]',
                    'input[value*="搜尋"]',
                ]

                for selector in button_selectors:
                    try:
                        query_button = self.driver.find_element(
                            By.CSS_SELECTOR, selector
                        )
                        query_button.click()
                        self.logger.info(f"✅ 已點擊查詢按鈕", operation="search")
                        time.sleep(Timeouts.QUERY_SUBMIT)
                        query_button_found = True
                        break
                    except Exception:
                        continue

                if not query_button_found:
                    self.logger.warning(f"⚠️ 未找到查詢按鈕，直接繼續流程", operation="search")
            else:
                self.logger.warning(f"⚠️ 未找到月份輸入框，可能不需要設定月份", operation="config")

            return True

        except Exception as e:
            # 使用診斷管理器捕獲日期設定異常
            diagnostic_report = self.diagnostic_manager.capture_exception(
                e,
                context={
                    "operation": "set_date_range",
                    "username": self.username,
                    "start_month": self.start_month,
                    "end_month": self.end_month,
                    "current_url": self.driver.current_url if self.driver else None,
                },
                capture_screenshot=True,
                capture_page_source=True,
                driver=self.driver,
            )

            self.logger.warning(
                f"⚠️ 月份範圍設定過程中出現問題，但繼續執行: {e}",
                operation="config",
                diagnostic_report=diagnostic_report,
            )
            return True  # 即使失敗也返回True，讓流程繼續

    def get_freight_records(self) -> List[Dict[str, Any]]:
        """搜尋運費相關數據 - 基於wedi_selenium_scraper.py的邏輯但適配運費"""
        assert self.driver is not None, "WebDriver must be initialized"
        self.logger.info(f"📊 搜尋運費數據...", operation="search")

        records: List[Dict[str, Any]] = []
        try:
            # 此時已經在datamain iframe中，直接搜尋數據
            self.logger.debug(f"🔍 分析當前頁面內容...")

            # 先搜尋表格中的發票數據（運費查詢結果為表格格式）
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            self.logger.info(f"   找到 {len(tables)} 個表格")

            for table_index, table in enumerate(tables):
                try:
                    rows = table.find_elements(By.TAG_NAME, "tr")
                    self.logger.info(f"   表格 {table_index + 1} 有 {len(rows)} 行")

                    # 詳細分析每一行的內容
                    for row_index, row in enumerate(rows):
                        try:
                            cells = row.find_elements(By.TAG_NAME, "td")
                            th_cells = row.find_elements(By.TAG_NAME, "th")
                            total_cells = len(cells) + len(th_cells)

                            if total_cells > 0:
                                self.logger.info(
                                    f"   行 {row_index + 1}: {len(cells)} 個 td, {len(th_cells)} 個 th"
                                )

                                # 檢查每個欄位的內容
                                all_cells = cells if cells else th_cells
                                for cell_index, cell in enumerate(all_cells):
                                    cell_text = cell.text.strip()
                                    if cell_text:
                                        self.logger.info(
                                            f"     欄位 {cell_index + 1}: '{cell_text}'"
                                        )

                                        # 檢查這個欄位是否包含發票號碼（英數字組合，長度 > 8）
                                        # 排除包含中文字符、特殊符號（如 - 後接中文）的客戶名稱
                                        is_invoice_like = (
                                            len(cell_text) > 8
                                            and any(c.isdigit() for c in cell_text)
                                            and any(c.isalpha() for c in cell_text)
                                            and cell_text not in ["發票號碼", "小計", "總計"]
                                        )

                                        # 排除包含中文字符的客戶名稱（如 5081794203-宥芯有限公）
                                        has_chinese = any(
                                            ord(c) >= 0x4E00 and ord(c) <= 0x9FFF
                                            for c in cell_text
                                        )

                                        # 排除明顯的客戶代碼格式（數字-中文公司名）
                                        is_customer_code = (
                                            "-" in cell_text and has_chinese
                                        )

                                        if (
                                            is_invoice_like
                                            and not has_chinese
                                            and not is_customer_code
                                        ):
                                            self.logger.debug(
                                                f"     🔍 可能的發票號碼: '{cell_text}'"
                                            )

                                            # 檢查是否有可點擊的連結
                                            invoice_link = None
                                            try:
                                                # 嘗試在該欄位中尋找連結
                                                invoice_link = cell.find_element(
                                                    By.TAG_NAME, "a"
                                                )
                                                self.logger.info(f"     ✅ 在欄位中找到連結")
                                            except Exception:
                                                # 如果該欄位沒有連結，嘗試整行是否可點擊
                                                try:
                                                    invoice_link = row.find_element(
                                                        By.TAG_NAME, "a"
                                                    )
                                                    self.logger.info(f"     ✅ 在整行中找到連結")
                                                except Exception:
                                                    # 如果沒有連結，使用整個 cell 作為點擊目標
                                                    invoice_link = cell
                                                    self.logger.warning(
                                                        f"     ⚠️ 沒有連結，使用欄位本身"
                                                    )

                                            if invoice_link:
                                                # 嘗試獲取發票日期（通常在前一個或後一個欄位）
                                                invoice_date = ""
                                                try:
                                                    # 檢查前後欄位是否有日期格式（8位數字）
                                                    for check_index in [
                                                        cell_index - 1,
                                                        cell_index + 1,
                                                    ]:
                                                        if (
                                                            0
                                                            <= check_index
                                                            < len(all_cells)
                                                        ):
                                                            check_text = all_cells[
                                                                check_index
                                                            ].text.strip()
                                                            if (
                                                                len(check_text) == 8
                                                                and check_text.isdigit()
                                                            ):
                                                                invoice_date = (
                                                                    check_text
                                                                )
                                                                break
                                                except Exception:
                                                    pass

                                                records.append(
                                                    {
                                                        "index": len(records) + 1,
                                                        "title": f"發票號碼: {cell_text}",
                                                        "invoice_no": cell_text,
                                                        "invoice_date": invoice_date,
                                                        "record_id": (
                                                            f"{invoice_date}_{cell_text}"
                                                            if invoice_date
                                                            else cell_text
                                                        ),
                                                        "link": invoice_link,
                                                    }
                                                )
                                                self.logger.info(
                                                    f"   ✅ 找到發票記錄: {cell_text} (日期: {invoice_date})"
                                                )

                        except Exception as row_e:
                            self.logger.warning(
                                f"   ⚠️ 處理行 {row_index + 1} 時出錯: {row_e}"
                            )
                            continue

                except Exception as table_e:
                    self.logger.warning(f"   ⚠️ 處理表格 {table_index + 1} 時出錯: {table_e}")
                    continue

            # 如果表格中沒有找到發票數據，嘗試搜尋連結
            if not records:
                self.logger.debug(f"🔍 表格中未找到發票數據，搜尋連結...", operation="search")
                all_links = self.driver.find_elements(By.TAG_NAME, "a")
                self.logger.info(f"   找到 {len(all_links)} 個連結")

                # 定義運費相關關鍵字

                # 定義排除關鍵字
                excluded_keywords = [
                    "語音取件",
                    "三節加價",
                    "系統公告",
                    "操作說明",
                    "維護通知",
                    "Home",
                    "首頁",
                    "登出",
                    "系統設定",
                    "有限公司",
                    "股份有限公司",
                    "企業",
                    "公司",
                    "-",  # 避免客戶代碼格式被識別
                ]

                for i, link in enumerate(all_links):
                    try:
                        link_text = link.text.strip()
                        if link_text:
                            # 檢查是否需要排除
                            should_exclude = any(
                                keyword in link_text for keyword in excluded_keywords
                            )

                            # 匹配運費相關項目或發票號碼格式
                            is_freight_record = (
                                ("運費" in link_text and "月結" in link_text)
                                or ("結帳資料" in link_text and "運費" in link_text)
                                or "(2-7)" in link_text
                                or (
                                    len(link_text) > 8
                                    and any(c.isdigit() for c in link_text)
                                    and any(c.isalpha() for c in link_text)
                                )
                            )

                            if is_freight_record and not should_exclude:
                                # 生成檔案ID
                                file_id = (
                                    link_text.replace(" ", "_")
                                    .replace("[", "")
                                    .replace("]", "")
                                    .replace("-", "_")
                                )
                                records.append(
                                    {
                                        "index": i + 1,
                                        "title": link_text,
                                        "record_id": file_id,
                                        "link": link,
                                    }
                                )
                                self.logger.info(f"   ✅ 找到運費記錄: {link_text}")
                            elif should_exclude:
                                self.logger.info(f"   ⏭️ 跳過排除項目: {link_text}")
                    except Exception:
                        continue

            self.logger.info(f"📊 總共找到 {len(records)} 筆運費相關記錄")
            return records

        except Exception as e:
            # 使用診斷管理器捕獲搜尋異常
            diagnostic_report = self.diagnostic_manager.capture_exception(
                e,
                context={
                    "operation": "get_freight_records",
                    "username": self.username,
                    "current_url": self.driver.current_url if self.driver else None,
                    "records_found": len(records),
                    "start_month": self.start_month,
                    "end_month": self.end_month,
                },
                capture_screenshot=True,
                capture_page_source=True,
                driver=self.driver,
            )

            self.logger.log_operation_failure(
                "搜尋運費數據", e, diagnostic_report=diagnostic_report
            )
            return records

    def download_excel_for_record(self, record: Dict[str, Any]) -> DownloadResult:
        """為特定運費記錄下載Excel檔案 - 修正stale element問題"""
        assert self.driver is not None, "WebDriver must be initialized"
        self.logger.info(
            f"📥 下載記錄 {record['record_id']} 的Excel檔案...", operation="download"
        )

        try:
            # 重新搜尋發票連結（避免 stale element reference）
            invoice_no = record["invoice_no"]
            self.logger.debug(f"🔍 重新搜尋發票號碼 {invoice_no} 的連結...", operation="search")

            # 提取純發票號碼（移除公司名稱部分）
            # 格式如：5081794203-宥芯有限公 -> 5081794203
            base_invoice_no = (
                invoice_no.split("-")[0] if "-" in invoice_no else invoice_no
            )
            self.logger.debug(f"   提取純發票號碼: {base_invoice_no}")

            found_link = None
            # 方法1：用完整發票號碼搜尋連結
            try:
                found_link = self.driver.find_element(
                    By.XPATH, f"//a[contains(text(), '{invoice_no}')]"
                )
                self.logger.info(f"✅ 通過完整文字內容找到連結")
            except Exception:
                # 方法2：用純發票號碼搜尋連結
                try:
                    found_link = self.driver.find_element(
                        By.XPATH, f"//a[contains(text(), '{base_invoice_no}')]"
                    )
                    self.logger.info(f"✅ 通過純發票號碼找到連結")
                except Exception:
                    # 方法3：通過 href 屬性搜尋（先完整後純號碼）
                    try:
                        found_link = self.driver.find_element(
                            By.XPATH, f"//a[contains(@href, '{invoice_no}')]"
                        )
                        self.logger.info(f"✅ 通過完整href屬性找到連結")
                    except Exception:
                        try:
                            found_link = self.driver.find_element(
                                By.XPATH, f"//a[contains(@href, '{base_invoice_no}')]"
                            )
                            self.logger.info(f"✅ 通過純發票號碼href屬性找到連結")
                        except Exception:
                            # 方法4：重新搜尋表格中的連結（更靈活的比對）
                            try:
                                tables = self.driver.find_elements(By.TAG_NAME, "table")
                                for table in tables:
                                    links = table.find_elements(By.TAG_NAME, "a")
                                    for link in links:
                                        link_text = link.text.strip()
                                        # 嘗試完整匹配或純號碼匹配
                                        if (
                                            invoice_no in link_text
                                            or base_invoice_no in link_text
                                            or link_text in invoice_no
                                        ):
                                            found_link = link
                                            self.logger.info(
                                                f"✅ 在表格中找到匹配連結: '{link_text}'"
                                            )
                                            break
                                    if found_link:
                                        break
                            except Exception as e:
                                self.logger.warning(
                                    f"⚠️ 重新搜尋連結失敗: {e}", operation="search"
                                )

            if found_link:
                # 使用JavaScript點擊避免元素遮蔽問題
                self.driver.execute_script("arguments[0].click();", found_link)
                self.logger.info(f"✅ 已點擊發票號碼 {invoice_no} 的連結")
                time.sleep(Timeouts.PAGE_LOAD)
            else:
                self.logger.warning(
                    f"⚠️ 找不到發票號碼 {invoice_no} 的連結，將嘗試 data-fileblob 方法",
                    operation="search",
                )

            downloaded_files = []
            record_id = record["record_id"]

            # 只有在沒有找到連結時才需要設定月份（此時在主查詢頁面）
            # 如果有連結並成功點擊，已經進入詳細頁面，可直接提取 data-fileblob
            if not found_link:
                self.logger.info(f"📅 在主查詢頁面填入查詢月份...", operation="search")
                try:
                    # 使用指定的月份範圍
                    if self.start_month and self.end_month:
                        start_month = self.start_month
                        end_month = self.end_month
                    else:
                        # 預設值（上個月）
                        start_month_str, end_month_str = self.get_default_date_range()
                        start_month = start_month_str
                        end_month = end_month_str

                    # 找到月份輸入框
                    date_inputs = self.driver.find_elements(
                        By.CSS_SELECTOR, 'input[type="text"]'
                    )
                    if len(date_inputs) >= 2:
                        # 填入開始月份
                        date_inputs[0].clear()
                        date_inputs[0].send_keys(start_month)
                        self.logger.info(f"✅ 已填入開始月份: {start_month}")

                        # 填入結束月份
                        date_inputs[1].clear()
                        date_inputs[1].send_keys(end_month)
                        self.logger.info(f"✅ 已填入結束月份: {end_month}")
                    elif len(date_inputs) >= 1:
                        # 只有一個月份輸入框，填入查詢月份
                        date_inputs[0].clear()
                        date_inputs[0].send_keys(start_month)
                        self.logger.info(
                            f"✅ 已填入查詢月份: {start_month}", operation="search"
                        )

                    # 點擊查詢按鈕
                    try:
                        query_button = self.driver.find_element(
                            By.CSS_SELECTOR, 'input[value*="查詢"]'
                        )
                        query_button.click()
                        self.logger.info(f"✅ 已點擊查詢按鈕", operation="search")
                        time.sleep(Timeouts.PAGE_LOAD)
                    except Exception:
                        self.logger.warning(f"⚠️ 未找到查詢按鈕，跳過此步驟", operation="search")

                except Exception as date_e:
                    self.logger.warning(
                        f"⚠️ 填入查詢月份失敗: {date_e}", error=str(date_e), operation="search"
                    )
            else:
                # 成功點擊連結進入詳細頁面，不需要設定月份，直接提取數據
                self.logger.info(f"✅ 已進入發票詳細頁面，準備提取數據...", operation="download")

            # 直接從頁面提取 data-fileblob 數據並生成 Excel
            try:
                self.logger.info(f"🚀 嘗試從頁面提取 data-fileblob 數據...")

                # 尋找包含 data-fileblob 屬性的按鈕
                fileblob_buttons = self.driver.find_elements(
                    By.CSS_SELECTOR, "button[data-fileblob]"
                )

                if not fileblob_buttons:
                    # 如果找不到，嘗試其他可能的選擇器
                    fileblob_buttons = self.driver.find_elements(
                        By.XPATH, "//*[@data-fileblob]"
                    )

                if fileblob_buttons:
                    self.logger.info(
                        f"✅ 找到 {len(fileblob_buttons)} 個包含 data-fileblob 的元素"
                    )

                    # 通常第一個就是我們要的匯出按鈕
                    fileblob_button = fileblob_buttons[0]
                    fileblob_data = fileblob_button.get_attribute("data-fileblob")

                    if fileblob_data:
                        self.logger.info(f"✅ 成功獲取 data-fileblob 數據")

                        try:
                            # 解析 JSON 數據
                            blob_json = json.loads(fileblob_data)
                            data_array = blob_json.get("data", [])
                            filename_base = blob_json.get("fileName", "Excel")
                            mime_type = blob_json.get(
                                "mimeType",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                            file_extension = blob_json.get("fileExtension", ".xlsx")

                            self.logger.info(f"📊 數據信息:")
                            self.logger.info(f"   檔名: {filename_base}{file_extension}")
                            self.logger.info(f"   MIME類型: {mime_type}")
                            self.logger.info(f"   數據行數: {len(data_array)}")

                            if data_array:
                                # 檢查實際數據筆數（排除表頭和彙總行如小計、總計）
                                summary_keywords = ["小計", "總計", "合計"]
                                data_row_count = 0
                                for row in data_array[1:]:  # 跳過表頭
                                    # 檢查該行是否為彙總行
                                    row_text = "".join(str(cell) for cell in row)
                                    is_summary_row = any(keyword in row_text for keyword in summary_keywords)
                                    if not is_summary_row:
                                        data_row_count += 1
                                
                                if data_row_count <= 0:
                                    self.logger.info("📭 查詢結果筆數為 0，跳過下載")
                                    return []

                                # 使用 openpyxl 創建 Excel 檔案
                                wb = openpyxl.Workbook()
                                ws = wb.active
                                assert (
                                    ws is not None
                                ), "Workbook active sheet should not be None"
                                ws.title = "運費明細"

                                # 將數據寫入工作表
                                for row_index, row_data in enumerate(data_array, 1):
                                    for col_index, cell_value in enumerate(row_data, 1):
                                        # 清理數據（移除HTML空格等）
                                        if isinstance(cell_value, str):
                                            cell_value = cell_value.replace(
                                                "&nbsp;", ""
                                            ).strip()

                                        ws.cell(
                                            row=row_index,
                                            column=col_index,
                                            value=cell_value,
                                        )

                                # 設定表頭樣式
                                if len(data_array) > 0:
                                    from openpyxl.styles import Font, PatternFill

                                    # 表頭加粗
                                    for col_index in range(1, len(data_array[0]) + 1):
                                        cell = ws.cell(row=1, column=col_index)
                                        cell.font = Font(bold=True)
                                        cell.fill = PatternFill(
                                            start_color="CCCCCC",
                                            end_color="CCCCCC",
                                            fill_type="solid",
                                        )

                                # 自動調整欄寬
                                from openpyxl.cell.cell import Cell

                                for column in ws.columns:
                                    max_length = 0
                                    # 取得第一個 Cell 的 column_letter (跳過 MergedCell)
                                    column_letter = None
                                    for cell in column:
                                        if isinstance(cell, Cell) and hasattr(
                                            cell, "column_letter"
                                        ):
                                            column_letter = cell.column_letter
                                            break

                                    if column_letter is None:
                                        continue

                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except Exception:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)  # 最大寬度限制
                                    ws.column_dimensions[
                                        column_letter
                                    ].width = adjusted_width

                                # 生成檔案名稱
                                # 優先使用 data-fileblob 中的檔案名，因為這是實際下載的內容
                                actual_invoice_info = ""
                                search_invoice_no = record.get("invoice_no", record_id)
                                search_invoice_date = record.get("invoice_date", "")

                                # 嘗試從 data-fileblob 的檔案名中提取實際發票資訊
                                if filename_base and filename_base != "Excel":
                                    # 如果 data-fileblob 有有用的檔案名，使用它
                                    actual_invoice_info = filename_base
                                    self.logger.info(
                                        f"📄 使用實際下載檔案名: {actual_invoice_info}"
                                    )
                                else:
                                    # 回退到搜尋時的發票號碼
                                    actual_invoice_info = search_invoice_no
                                    self.logger.info(
                                        f"📄 回退使用搜尋發票號碼: {actual_invoice_info}"
                                    )

                                # 生成最終檔案名
                                if search_invoice_date:
                                    filename = f"運費發票明細_{self.username}_{search_invoice_date}_{actual_invoice_info}.xlsx"
                                else:
                                    filename = f"運費發票明細_{self.username}_{actual_invoice_info}.xlsx"

                                # 記錄檔案命名邏輯以供調試
                                self.logger.info(f"🏷️ 檔案命名資訊:")
                                self.logger.info(f"   搜尋發票號碼: {search_invoice_no}")
                                self.logger.info(f"   實際檔案內容: {actual_invoice_info}")
                                self.logger.info(f"   最終檔案名: {filename}")

                                # 檢查檔案是否已下載
                                exists, existing_path = self.is_file_downloaded(filename)
                                if exists:
                                    wb.close()  # 關閉未使用的 workbook
                                    self.logger.info(
                                        f"⏭️ 檔案已存在，跳過生成: {filename}",
                                        location=str(existing_path)
                                    )
                                    return [str(existing_path)]
                                
                                # 保存檔案
                                file_path = self.download_dir / filename

                                # 確保下載目錄存在且可寫入（提供詳細診斷訊息）
                                self.ensure_directory_writable(self.download_dir)

                                wb.save(file_path)
                                wb.close()

                                downloaded_files = [str(file_path)]
                                self.logger.info(
                                    f"✅ 成功從 data-fileblob 生成 Excel: {filename}"
                                )
                                self.logger.info(
                                    f"📁 檔案大小: {file_path.stat().st_size:,} bytes"
                                )
                                self.logger.info(
                                    f"📋 數據行數: {len(data_array)} 行，"
                                    f"欄數: {len(data_array[0]) if data_array else 0} 欄"
                                )

                                return downloaded_files

                            else:
                                self.logger.error(f"❌ data-fileblob 中沒有找到數據陣列")
                                # 如果連結也找不到，這是完全失敗
                                if not found_link:
                                    self.logger.error(
                                        f"❌ 發票號碼 {invoice_no} 無法下載：連結未找到且 data-fileblob 無數據"
                                    )
                                return []

                        except json.JSONDecodeError as json_e:
                            self.logger.error(
                                f"❌ 解析 data-fileblob JSON 失敗: {json_e}",
                                error=str(json_e),
                            )
                            self.logger.info(f"   原始數據前500字元: {fileblob_data[:500]}")
                            return []

                        except Exception as excel_e:
                            self.logger.error(
                                f"❌ 生成 Excel 檔案失敗: {excel_e}", error=str(excel_e)
                            )
                            return []

                    else:
                        self.logger.error(f"❌ data-fileblob 屬性為空")
                        return []

                else:
                    self.logger.warning(f"⚠️ 未找到包含 data-fileblob 的元素")
                    # 不拋出異常，返回空列表讓程式優雅處理
                    return []

            except Exception as blob_e:
                self.logger.error(f"❌ data-fileblob 提取失敗: {blob_e}", error=str(blob_e))
                # 如果連結也找不到，說明兩種方法都失敗了
                if not found_link:
                    self.logger.error(
                        f"❌ 發票號碼 {invoice_no} 無法下載：連結未找到且 data-fileblob 提取失敗"
                    )
                else:
                    self.logger.info(f"🔄 程式無法提取數據，請檢查頁面是否正確載入")
                return []

        except Exception as e:
            # 使用診斷管理器捕獲下載異常
            diagnostic_report = self.diagnostic_manager.capture_exception(
                e,
                context={
                    "operation": "download_excel_for_record",
                    "username": self.username,
                    "record": record,
                    "current_url": self.driver.current_url if self.driver else None,
                    "start_month": self.start_month,
                    "end_month": self.end_month,
                },
                capture_screenshot=True,
                capture_page_source=True,
                driver=self.driver,
            )

            self.logger.log_operation_failure(
                "下載記錄", e, diagnostic_report=diagnostic_report
            )
            return []

    def run_full_process(self) -> List[str]:
        """執行完整的自動化流程"""
        all_downloads: DownloadResult = []
        records: List[Dict[str, Any]] = []

        try:
            self.logger.info("=" * 60)
            self.logger.info(
                f"🚛 開始執行 WEDI 運費查詢流程 (帳號: {self.username})", operation="search"
            )
            self.logger.info("=" * 60)

            # 1. 瀏覽器已在 __init__ 中初始化

            # 2. 登入
            login_success = self.login()
            if not login_success:
                self.logger.log_operation_failure("登入", Exception("登入失敗"))
                return []

            # 3. 導航到查詢頁面（基礎導航）
            nav_success = self.navigate_to_query()
            if not nav_success:
                self.logger.log_operation_failure("基礎導航", Exception("導航失敗"))
                return []

            # 4. 導航到運費查詢頁面
            freight_nav_success = self.navigate_to_freight_page()
            if not freight_nav_success:
                self.logger.log_operation_failure("運費頁面導航", Exception("導航失敗"))
                return []

            # 5. 設定月份範圍
            self.set_date_range()

            # 6. 獲取運費記錄
            records = self.get_freight_records()

            if not records:
                self.logger.warning(f"⚠️ 帳號 {self.username} 沒有找到運費記錄")
                return []

            # 7. 下載每個記錄的Excel檔案
            for record in records:
                try:
                    downloads = self.download_excel_for_record(record)
                    all_downloads.extend(downloads)
                except Exception as download_e:
                    self.logger.warning(
                        f"⚠️ 帳號 {self.username} 下載記錄 "
                        f"{record.get('record_id', 'unknown')} 失敗: {download_e}",
                        operation="download",
                    )
                    continue

            self.logger.info(f"🎉 帳號 {self.username} 自動化流程完成！")
            return all_downloads

        except Exception as e:
            self.logger.info(f"💥 帳號 {self.username} 流程執行失敗: {e}", error=str(e))
            return all_downloads
        finally:
            self.close()


def main():
    """主程式入口"""
    from src.core.logging_config import get_logger

    # 設置主程式日誌器
    logger = get_logger("freight_scraper_main")

    parser = argparse.ArgumentParser(description="WEDI 運費查詢自動下載工具")
    parser.add_argument("--headless", action="store_true", help="使用無頭模式")
    parser.add_argument("--start-month", type=str, help="開始月份 (格式: YYYYMM，例如: 202411)")
    parser.add_argument("--end-month", type=str, help="結束月份 (格式: YYYYMM，例如: 202412)")

    args = parser.parse_args()

    # 月份參數驗證和處理
    start_month = None
    end_month = None

    if args.start_month:
        try:
            # 驗證月份格式
            year = int(args.start_month[:4])
            month = int(args.start_month[4:6])
            if not (1 <= month <= 12):
                raise ValueError("月份必須在01-12之間")
            start_month = args.start_month
        except (ValueError, IndexError) as e:
            logger.error(f"⛔ 開始月份格式錯誤: {e}")
            logger.info("💡 月份格式應為 YYYYMM，例如: 202411")
            return 1

    if args.end_month:
        try:
            # 驗證月份格式
            year = int(args.end_month[:4])
            month = int(args.end_month[4:6])
            if not (1 <= month <= 12):
                raise ValueError("月份必須在01-12之間")
            end_month = args.end_month
        except (ValueError, IndexError) as e:
            logger.error(f"⛔ 結束月份格式錯誤: {e}")
            logger.info("💡 月份格式應為 YYYYMM，例如: 202412")
            return 1

    # 如果沒有指定月份，使用預設值
    if not start_month:
        today = datetime.now()
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year

        start_month = f"{prev_year:04d}{prev_month:02d}"
        end_month = start_month  # 預設查詢單一月份

        logger.info(
            f"📅 查詢月份範圍: {start_month} ~ {end_month} (預設上個月)",
            start_month=start_month,
            end_month=end_month,
        )
    elif not end_month:
        end_month = start_month  # 如果只指定開始月份，結束月份使用相同值
        logger.info(
            f"📅 查詢月份範圍: {start_month} ~ {end_month}",
            start_month=start_month,
            end_month=end_month,
        )
    else:
        logger.info(
            f"📅 查詢月份範圍: {start_month} ~ {end_month}",
            start_month=start_month,
            end_month=end_month,
        )

    try:
        # 使用多帳號管理器
        logger.info(f"🚛 WEDI 運費查詢自動下載工具", operation="download")

        manager = MultiAccountManager("accounts.json")
        manager.run_all_accounts(
            scraper_class=FreightScraper,
            headless_override=args.headless if args.headless else None,
            start_month=start_month,
            end_month=end_month,
        )

        return 0

    except (FileNotFoundError, ValueError, RuntimeError) as e:
        logger.error(f"⛔ 錯誤: {e}")
        return 1
    except KeyboardInterrupt:
        logger.warning("\n⛔ 使用者中斷執行")
        return 1
    except Exception as e:
        logger.error(f"⛔ 未知錯誤: {e}")
        return 1


if __name__ == "__main__":
    main()
